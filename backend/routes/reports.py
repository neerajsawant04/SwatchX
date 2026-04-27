from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt
from bson import ObjectId
import datetime, io, os
from database import complaints_col, reports_col

reports_bp = Blueprint("reports", __name__)

def _admin_guard():
    if get_jwt().get("role") != "admin":
        return jsonify({"error": "Admin only"}), 403
    return None

def _s(doc): doc["_id"] = str(doc.get("_id", "")); return doc

def _fetch_complaints(range_type, start_date=None, end_date=None, limit=None):
    now   = datetime.datetime.utcnow()
    query = {}
    if range_type == "last10":
        docs = list(complaints_col().find({}, sort=[("timestamp", -1)], limit=10))
    elif range_type == "last30":
        since = now - datetime.timedelta(days=30)
        docs  = list(complaints_col().find({"timestamp": {"$gte": since}}, sort=[("timestamp", -1)]))
    elif range_type == "custom" and start_date and end_date:
        docs = list(complaints_col().find({
            "timestamp": {"$gte": start_date, "$lte": end_date}
        }, sort=[("timestamp", -1)]))
    else:
        docs = list(complaints_col().find({}, sort=[("timestamp", -1)], limit=50))
    return [_s(d) for d in docs]

@reports_bp.route("/preview", methods=["GET"])
@jwt_required()
def preview():
    g = _admin_guard()
    if g: return g
    range_type = request.args.get("range", "last10")
    docs       = _fetch_complaints(range_type)
    return jsonify({"complaints": docs, "count": len(docs), "range": range_type}), 200

@reports_bp.route("/export/excel", methods=["GET"])
@jwt_required()
def export_excel():
    g = _admin_guard()
    if g: return g
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    range_type = request.args.get("range", "last30")
    docs       = _fetch_complaints(range_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WasteGuard Report"

    # Styling
    header_fill  = PatternFill("solid", fgColor="1A3C1A")
    header_font  = Font(bold=True, color="A3E635", size=11)
    alt_fill     = PatternFill("solid", fgColor="0F1A0F")
    center       = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin", color="2A3C2A"),
        right=Side(style="thin", color="2A3C2A"),
        top=Side(style="thin", color="2A3C2A"),
        bottom=Side(style="thin", color="2A3C2A")
    )

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value     = f"WasteGuard — Complaint Report ({range_type.upper()})"
    title_cell.font      = Font(bold=True, color="A3E635", size=14)
    title_cell.fill      = PatternFill("solid", fgColor="0A0F0A")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # Sub-title
    ws.merge_cells("A2:J2")
    ws["A2"].value     = f"Generated: {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Total: {len(docs)} complaints"
    ws["A2"].font      = Font(italic=True, color="86908A", size=10)
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 20

    headers = ["#", "Complaint No.", "User Email", "Waste Type",
               "Environmental Impact", "Detected Materials", "Agency Email", "Pincode",
               "Status", "Timestamp", "Resolved At"]

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border
    ws.row_dimensions[3].height = 22

    for row_idx, doc in enumerate(docs, 4):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="142014")
        row_data = [
            row_idx - 3,
            doc.get("complaintNumber", ""),
            doc.get("userEmail", ""),
            doc.get("wasteType", ""),
            doc.get("environmentalImpact"),
            doc.get("materialsList", ""),
            doc.get("agencyEmail", ""),
            doc.get("pincode", ""),
            doc.get("status", ""),
            str(doc.get("timestamp", ""))[:19].replace("T", " "),
            str(doc.get("resolvedAt", "") or "—")[:19]
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill   = fill
            cell.border = thin_border
            cell.font   = Font(color="D1FAE5" if doc.get("status")=="Cleaned" else "F1F5F0", size=10)
            if col in (1, 8): cell.alignment = center
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    col_widths = [5, 22, 30, 20, 40, 30, 12, 12, 22, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Save
    output = io.BytesIO()
    wb.save(output); output.seek(0)

    fname = f"WasteGuard_Report_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)

@reports_bp.route("/export/pdf", methods=["GET"])
@jwt_required()
def export_pdf():
    g = _admin_guard()
    if g: return g
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: pip install reportlab"}), 500

    range_type = request.args.get("range", "last30")
    docs       = _fetch_complaints(range_type)
    output     = io.BytesIO()

    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)

    BG       = colors.HexColor("#0A0F0A")
    ACCENT   = colors.HexColor("#A3E635")
    SURFACE  = colors.HexColor("#142014")
    TEXT     = colors.HexColor("#F1F5F0")
    MUTED    = colors.HexColor("#86908A")
    GREEN    = colors.HexColor("#4ADE80")
    RED      = colors.HexColor("#F87171")
    YELLOW   = colors.HexColor("#FBBF24")

    styles = getSampleStyleSheet()
    title_style  = ParagraphStyle("title",  fontSize=18, textColor=ACCENT, fontName="Helvetica-Bold", spaceAfter=4)
    sub_style    = ParagraphStyle("sub",    fontSize=9,  textColor=MUTED,  fontName="Helvetica", spaceAfter=12)
    wrap_style   = ParagraphStyle("wrap",   fontSize=7.5, textColor=TEXT,  fontName="Helvetica")

    story = [
        Paragraph("♻ WasteGuard — Complaint Report", title_style),
        Paragraph(f"Range: {range_type.upper()} | Generated: {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Total: {len(docs)}", sub_style),
        Spacer(1, 0.3*cm)
    ]

    headers = [["#", "Complaint No.", "User Email", "Waste Type", "Detected Materials",
                 "Env. Impact", "Agency", "Pincode", "Status", "Timestamp"]]
    table_data = headers[:]
    for i, doc_item in enumerate(docs, 1):
        status     = doc_item.get("status", "Pending")
        impact_txt = (doc_item.get("environmentalImpact","") or "")[:55]
        table_data.append([
            str(i),
            doc_item.get("complaintNumber",""),
            doc_item.get("userEmail",""),
            doc_item.get("wasteType",""),
            Paragraph(doc_item.get("environmentalImpact","")[:200], wrap_style),
            Paragraph(doc_item.get("materialsList",""), wrap_style),
            doc_item.get("agencyEmail",""),
            doc_item.get("pincode",""),
            status,
            str(doc_item.get("timestamp",""))[:16]
        ])

    col_widths_pt = [1*cm, 3.5*cm, 4.5*cm, 3*cm, 5.5*cm, 4.5*cm, 2*cm, 2.2*cm, 3.5*cm]
    tbl = Table(table_data, colWidths=col_widths_pt, repeatRows=1)

    style = TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), ACCENT),
        ("TEXTCOLOR",   (0,0), (-1,0), BG),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 8),
        ("ALIGN",       (0,0), (-1,0), "CENTER"),
        ("BACKGROUND",  (0,1), (-1,-1), SURFACE),
        ("TEXTCOLOR",   (0,1), (-1,-1), TEXT),
        ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",    (0,1), (-1,-1), 7.5),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [SURFACE, colors.HexColor("#0F1A0F")]),
        ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#1A2E1A")),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("PADDING",     (0,0), (-1,-1), 4),
    ])
    # Colour-code status cells
    for row_idx, row in enumerate(table_data[1:], 1):
        st = row[7] if isinstance(row[7], str) else ""
        c  = GREEN if st == "Cleaned" else (RED if st == "Rejected" else YELLOW)
        style.add("TEXTCOLOR", (7, row_idx), (7, row_idx), c)
        style.add("FONTNAME",  (7, row_idx), (7, row_idx), "Helvetica-Bold")

    tbl.setStyle(style)
    story.append(tbl)
    doc.build(story)
    output.seek(0)

    fname = f"WasteGuard_Report_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(output, mimetype="application/pdf",
                     as_attachment=True, download_name=fname)
